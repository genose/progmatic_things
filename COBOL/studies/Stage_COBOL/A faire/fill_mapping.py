#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fill_mapping.py — Remplit Mapping_D02_EXTCDE_CRMCSP1.xlsx
avec le dictionnaire de données complet (T1, T2, T3).

Usage :
    /usr/local/bin/python3 fill_mapping.py

Prérequis : openpyxl  (pip install openpyxl ou utiliser extract_mapping.sh)
"""

import openpyxl
import os

XLSX = os.path.join(os.path.dirname(__file__), "Mapping_D02_EXTCDE_CRMCSP1.xlsx")

wb = openpyxl.load_workbook(XLSX)
ws = wb.active

# ─────────────────────────────────────────────────────────────────────────────
# Helper : efface les lignes de données (3 à 91) dans les 7 colonnes d'un tableau
# ─────────────────────────────────────────────────────────────────────────────
def clear_table(start_col, start_row=3, end_row=91):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, start_col + 7):
            ws.cell(row=r, column=c).value = None

# ─────────────────────────────────────────────────────────────────────────────
# Helper : écrit un tableau de tuples (N°, Description, Nom, Format, Taille,
#          Commentaire, Exemple) à partir de start_col et start_row.
# ─────────────────────────────────────────────────────────────────────────────
def write_table(rows, start_col, start_row=3):
    for i, row in enumerate(rows):
        r = start_row + i
        n, desc, nom, fmt, taille, commentaire, exemple = row
        ws.cell(row=r, column=start_col + 0).value = n
        ws.cell(row=r, column=start_col + 1).value = desc
        ws.cell(row=r, column=start_col + 2).value = nom
        ws.cell(row=r, column=start_col + 3).value = fmt
        ws.cell(row=r, column=start_col + 4).value = taille
        ws.cell(row=r, column=start_col + 5).value = commentaire
        ws.cell(row=r, column=start_col + 6).value = exemple

# ═════════════════════════════════════════════════════════════════════════════
# TABLE 1 — Fichier positionnel D02_EXTCDE_CRMCSP1  (colonnes A-G, indices 1-7)
# 38 champs, positions absolues 1-197 dans l'enregistrement de 197 caractères.
# ═════════════════════════════════════════════════════════════════════════════
T1 = [
    (1, "Type d'enregistrement", "TYPMES", "CHAR", 6,
     "Discriminant du type de l'enregistrement. Constante écrite par MOVE '...' TO ENR-TYPMES avant chaque WRITE. "
     "Valeurs possibles : DEBCDE (début de fichier, 1 occurrence), REFCDE (en-tête d'une commande), "
     "LINTXT (commentaire BL ou RT, 1 par commande), TXTCDE (horaire QU00013 et/ou contact QU000132, conditionnel), "
     "LINCDE (1 par entrée CODENR='3' de TABLE-CDL), FINCDE (fin de groupe laboratoire, routé par PREV-CODDEP), "
     "FINMES (fin de fichier, 1 occurrence). Positions 1-6 de l'enregistrement de 197 caractères.",
     "LINCDE"),

    (2, "Séparateur positionnel fixe", "FILLER", "CHAR", 1,
     "Zone filler entre TYPMES (pos. 1-6) et le corps de l'enregistrement (pos. 8-197). Toujours espace. "
     "Alimenté par MOVE SPACES TO ENR-FILLER avant chaque WRITE. "
     "N'est jamais issu d'une donnée métier. Position absolue 7.",
     "(espace)"),

    (3, "Émetteur EDI", "DEBCDE-EMETEUR", "CHAR", 35,
     "Identifiant de l'émetteur du flux EDI. Constante littérale '183 CSP' complétée à 35 caractères par des espaces "
     "(padding droit COBOL). Alimenté par MOVE '183 CSP' ... dans WRITE-DEBCDE. "
     "Présent uniquement dans l'enregistrement DEBCDE, écrit une seule fois en tête de TRANSCO et TRANSMO. "
     "Positions absolues 8-42.",
     "183 CSP                            "),

    (4, "Récepteur EDI", "DEBCDE-RECEPTE", "CHAR", 35,
     "Identifiant du récepteur du flux EDI. Même valeur que DEBCDE-EMETEUR : constante '183 CSP' complétée à 35 "
     "caractères. Symétrique de l'émetteur dans le protocole EDI. Positions absolues 43-77.",
     "183 CSP                            "),

    (5, "Indicateur production/test", "DEBCDE-TEST", "CHAR", 1,
     "Flag permettant au récepteur de distinguer un flux de production d'un flux de test. "
     "Constante 'P' = production. Alimenté par MOVE 'P' dans WRITE-DEBCDE. Position absolue 78.",
     "P"),

    (6, "Rembourrage DEBCDE", "DEBCDE-FILLER", "CHAR", 100,
     "Zone réservée non utilisée dans l'enregistrement DEBCDE. Toujours espaces : mis à zéro/espaces par "
     "INITIALIZE ENREG-TRANSMIT avant chaque WRITE, jamais réalimenté. Positions absolues 79-178. "
     "Partagée avec les corps REDEFINES des autres types de record.",
     "(100 espaces)"),

    (7, "Rembourrage fin DEBCDE", "FILLER", "CHAR", 19,
     "Zone de rembourrage de fin d'enregistrement DEBCDE. Toujours espaces (INITIALIZE). "
     "Positions absolues 179-197. Ces octets correspondent dans le plan mémoire REDEFINES "
     "aux octets 172-190 du corps partagé avec REFCDE, LINTXT, etc.",
     "(19 espaces)"),

    (8, "Date de commande (YYMMDD)", "REFCDE-DATCDE", "CHAR", 6,
     "Date de création de la commande au format YYMMDD (siècle ignoré, 6 caractères). "
     "Calcul COBOL : CDE.DATCDE (DATE VMS binaire) → SYS$ASCTIM → W-DATE-ASCII 'DD-MON-YYYY HH:MI:SS.CC' "
     "→ recomposition YYYYMMDD dans FIC-DATCDE → tronqué aux positions 3-8 via FIC-DATCDE(3:6). "
     "Ex : '20260815' → '260815'. INSPECT remplace espaces par '0'. Positions absolues 8-13 dans REFCDE.",
     "260815"),

    (9, "Référence laboratoire de la commande", "REFCDE-CDELAB", "CHAR", 22,
     "Numéro de commande attribué par le laboratoire, distinct du numéro interne CSP. "
     "Source : D.CDE.CDELAB lu directement via CURCDE ou CURCDE_R. "
     "MOVE CDELAB IN CDE TO ENR-REFCDE-CDELAB. "
     "Permet au laboratoire de retrouver la commande dans son propre système. "
     "Positions absolues 15-36 dans REFCDE.",
     "LAB2024-00123456789   "),

    (10, "Référence interne commande CSP", "REFCDE-REFCDE", "CHAR", 35,
     "Référence interne de la commande telle qu'elle existe dans BD_DEPOT.CDE. "
     "Lu dans CURCDE/CURCDE_R via la variable hôte SQL :w-refcde (PIC X(35), déclarée ligne 367 du source). "
     "MOVE W-REFCDE TO ENR-REFCDE-REFCDE. Positions absolues 38-72 dans REFCDE.",
     "CSP-REF-2024-001234567890123456789"),

    (11, "Type de commande", "REFCDE-TYPCDE", "CHAR", 2,
     "Code classifiant le type de la commande. Source : D.CDE.TYPCDE lu via CURCDE/CURCDE_R. "
     "MOVE TYPCDE IN CDE TO ENR-REFCDE-TYPCDE. "
     "Exemples : 'ST' = standard, 'RA' = reste-à-livrer, 'GR' = gratuit. "
     "Positions absolues 74-75 dans REFCDE.",
     "ST"),

    (12, "Code client livré CSP", "REFCDE-CLICSP", "CHAR", 6,
     "Code numérique du client destinataire de la livraison (PIC 9(06)), affiché sur 6 chiffres avec zéros à gauche. "
     "Source : D.CDE.CLICSP. MOVE CLICSP IN CDE TO ENR-REFCDE-CLICSP. "
     "Ce même champ sert aussi de source pour REFCDE-NOMLIV (voir ligne suivante). "
     "Positions absolues 77-82 dans REFCDE.",
     "001234"),

    (13, "Décalage CLICSP vers champ NOMLIV", "REFCDE-NOMLIV", "CHAR", 35,
     "Champ nommé NOMLIV mais ne contenant PAS un nom de livraison. Astuce COBOL délibérée : "
     "(1) MOVE CLICSP IN CDE TO ENR-REFCDE-NOMLIV → ex '001234' + 29 espaces dans 35 chars ; "
     "(2) MOVE ENR-REFCDE-NOMLIV(4:6) TO ENR-REFCDE-NOMLIV → supprime les 3 premiers chiffres "
     "→ '234' + 32 espaces. Résultat : le récepteur reçoit les 3 derniers chiffres de CLICSP décalés à gauche. "
     "Comportement volontaire à reproduire EXACTEMENT tel quel en Java (voir §8 du guide TDD). "
     "Positions absolues 84-118 dans REFCDE.",
     "234                                "),

    (14, "Filler interne REFCDE", "REFCDE-FILLER", "CHAR", 1,
     "Zone filler interne à l'enregistrement REFCDE, positionnée entre NOMLIV et CODOPE. "
     "MOVE SPACES TO ENR-REFCDE-FILLER. Toujours espace. Position absolue 119 dans REFCDE.",
     "(espace)"),

    (15, "Code opérateur de saisie", "REFCDE-CODOPE", "CHAR", 3,
     "Identifiant de l'opérateur ou du système ayant créé ou modifié la commande dans BD_DEPOT. "
     "Source : D.CDE.CODOPE. MOVE CODOPE IN CDE TO ENR-REFCDE-CODOPE. "
     "Positions absolues 120-122 dans REFCDE.",
     "OPE"),

    (16, "Numéro de document (BD_PDF)", "REFCDE-NUMDOC", "CHAR", 10,
     "Numéro de document issu de la base BD_PDF (alias SQL P), table DOCENT, colonne NUMDOC. "
     "Recherché par le paragraphe RECH-NUMDOC : SELECT NUMDOC FROM P.DOCENT WHERE REFDOC STARTING WITH :W-NUMCDE. "
     "Si SQLSTATE = '02000' (non trouvé) : MOVE SPACES TO ENR-REFCDE-NUMDOC (10 espaces). "
     "Si trouvé : MOVE W-NUMDOC TO ENR-REFCDE-NUMDOC. Positions absolues 124-133 dans REFCDE.",
     "FR2026001234"),

    (17, "Indicateur commande saisie manuellement", "REFCDE-CDESAISIE", "CHAR", 1,
     "Indicateur signalant si la commande a été saisie manuellement ou importée automatiquement. "
     "Source : D.CDE.CDESAISIE. MOVE CDESAISIE IN CDE TO ENR-REFCDE-CDESAISIE. "
     "Valeurs typiques : 'O' = saisie manuelle, 'N' = import automatique. "
     "Position absolue 135 dans REFCDE.",
     "O"),

    (18, "Date/heure de réception (ASCII VMS 23 chars)", "REFCDE-DATREC", "CHAR", 23,
     "Date et heure de réception de la commande au format ASCII VMS complet : "
     "'DD-MON-YYYY HH:MI:SS.CC' (23 caractères). "
     "Source : D.CDE.DATREC converti par CALL SYS$ASCTIM → W-DATE-ASCII, "
     "puis MOVE W-DATE-ASCII TO FIC-DATREC. "
     "IMPORTANT — différence selon le curseur : CURCDE lit DATREC avec indicateur NULL "
     "(SQLSTATE '22002' = succès) ; CURCDE_R lit DATREC SANS indicateur NULL (comportement B05 du guide TDD). "
     "Positions absolues 137-159 dans REFCDE.",
     "15-AUG-2026 09:23:41.12"),

    (19, "Mode trafic transport", "REFCDE-TRAFIC", "CHAR", 1,
     "Mode de transport de la livraison. Source : D.CDE.TRAFIC. "
     "Règle de substitution COBOL : IF TRAFIC IN CDE = SPACES THEN MOVE 'N' TO ENR-REFCDE-TRAFIC "
     "ELSE MOVE TRAFIC IN CDE TO ENR-REFCDE-TRAFIC. "
     "Valeurs : 'E' = express, 'N' = normal (ou substitution si vide). "
     "Position absolue 161 dans REFCDE.",
     "N"),

    (20, "Date édition bon de préparation (YYYYMMDD)", "REFCDE-DATEBP", "CHAR", 8,
     "Date d'édition du bon de préparation au format YYYYMMDD (8 caractères). "
     "Source : D.CDE.DATEBP (DATE VMS binaire), indicateur NULL IDATEBP. "
     "Règle NULL : SI IDATEBP = -1 ALORS FIC-DATEBP = PARAM-DATE-TXT (date du jour en YYYYMMDD). "
     "Sinon : DATEBP → SYS$ASCTIM → recomposition YYYYMMDD (même algorithme que DATCDE). "
     "MOVE FIC-DATEBP TO ENR-REFCDE-DATEBP. Positions absolues 163-170 dans REFCDE.",
     "20260815"),

    (21, "Code saisie (opérateur web)", "REFCDE-CODSAI", "CHAR", 7,
     "Identifiant de l'opérateur web ou de l'interface de commande en ligne. "
     "Source : D.CDE.CODSAI. MOVE CODSAI IN CDE TO ENR-REFCDE-CODSAI. "
     "Positions absolues 172-178 dans REFCDE.",
     "WEBPDA1"),

    (22, "Code sous-traitant préparation", "REFCDE-CODSTR", "CHAR", 4,
     "Code du prestataire externe chargé de la préparation physique des colis. "
     "Source : D.CDE.CODSTR. MOVE CODSTR IN CDE TO ENR-REFCDE-CODSTR. "
     "Positions absolues 180-183 dans REFCDE.",
     "EXPL"),

    (23, "Code structure d'expédition", "REFCDE-STREXP", "CHAR", 4,
     "Code de la structure d'expédition (transporteur sous-traitant final). "
     "Source : D.CDE.STREXP. MOVE STREXP IN CDE TO ENR-REFCDE-STREXP. "
     "Positions absolues 185-188 dans REFCDE.",
     "CHRO"),

    (24, "Code représentant commercial", "REFCDE-CODREP", "CHAR", 8,
     "Code du représentant commercial affecté à la commande. "
     "Source : D.CDE.CODREP. MOVE CODREP IN CDE TO ENR-REFCDE-CODREP. "
     "Positions absolues 190-197 dans REFCDE.",
     "REP00123"),

    (25, "Commentaire BL ou RT (114 chars)", "LINTXT-MESSAGE", "CHAR", 114,
     "Texte du commentaire de livraison attaché à la commande. "
     "Source : D.MES.COMMENT (variable W-COMMENT PIC X(35)), lu par RECHERCHE-MES-BL-RT. "
     "Ordre de priorité : SELECT COMMENT FROM D.MES WHERE CODMES=CDE.CODMES AND TYPDOC='BL' en premier ; "
     "si absent, premier TYPDOC='RT'. Si aucun message : W-COMMENT = SPACES. "
     "MOVE W-COMMENT TO ENR-LINTXT-MESSAGE (padding espaces jusqu'à 114 chars). "
     "Un enregistrement LINTXT est toujours écrit (même si vide). "
     "Positions absolues 8-121 dans LINTXT.",
     "Livraison avant 12h - quai B                              "),

    (26, "Texte horaire ou contact livraison (80 chars)", "TXTCDE-MESSAGE", "CHAR", 80,
     "Texte libre issu de D.CDL.LIBELL pour les articles spéciaux de la commande. "
     "Article QU00013 → W-LIB-HORAIRE → enregistrement TXTCDE avec TYPDOC='RT'. "
     "Article QU000132 → W-LIB-CONTACT → enregistrement TXTCDE avec TYPDOC='BL'. "
     "SELECT LIBELL FROM D.CDL WHERE NUMCDE=:W-NUMCDE AND NUMRAL=:W-NUMRAL AND CODART='QU00013'. "
     "Enregistrement TXTCDE absent si l'article spécial n'existe pas sur la commande. "
     "Positions absolues 8-87 dans TXTCDE.",
     "Livraison 8h-12h lundi-vendredi - Contact : M. Dupont 06.12.34.56.78     "),

    (27, "Type document TXTCDE", "TXTCDE-TYPDOC", "CHAR", 2,
     "Discriminant du sous-type de l'enregistrement TXTCDE. "
     "Constante : 'RT' pour l'horaire de livraison (article QU00013, WRITE-TXTHOR) ; "
     "'BL' pour le contact livraison (article QU000132, WRITE-TXTCON). "
     "Positions absolues 89-90 dans TXTCDE.",
     "RT"),

    (28, "Code article CSP", "LINCDE-CODART", "CHAR", 10,
     "Code article CSP de la ligne de commande. "
     "Issu de TCDL-CODART(IND-CDL) dans TABLE-CDL (OCCURS 1000), entrée CODENR='3'. "
     "Source primaire : D.CDL.CODART lu par CURCDL. "
     "Un enregistrement LINCDE est généré par entrée CODENR='3'. "
     "MOVE TCDL-CODART(IND-CDL) TO ENR-LINCDE-CODART. "
     "Positions absolues 8-17 dans LINCDE.",
     "3400936750"),

    (29, "Quantité livrée cumulée (champ nommé QTCCDE)", "LINCDE-QTCCDE", "BIGINT", 7,
     "Malgré son nom QTCCDE (quantité commandée), ce champ contient la QUANTITÉ LIVRÉE cumulée. "
     "Source : TCDL-QTELIV(IND-CDL) = somme de D.CDL.QTLCDE pour toutes les lignes du groupe (NUMLIG, CODART). "
     "MOVE TCDL-QTELIV(IND-CDL) TO ENR-LINCDE-QTCCDE (commentaire source ligne 1831). "
     "Décision métier délibérée héritée d'une version antérieure. Reproduire tel quel en Java. "
     "PIC 9(07), 7 chiffres avec zéros à gauche. Positions absolues 19-25 dans LINCDE.",
     "0000144"),

    (30, "Quantité gratuite cumulée", "LINCDE-QTCGRT", "BIGINT", 7,
     "Quantité gratuite cumulée du groupe (NUMLIG, CODART). "
     "Source : TCDL-QTCGRT(IND-CDL) = somme de (D.CDL.QTCGRT + D.CDL.QTCECH) pour toutes les lignes du groupe. "
     "QTCECH (quantité échangée) est incluse dans le gratuit total "
     "(règle COBOL ADD QTCGRT QTCECH TO ...). "
     "PIC 9(07), 7 chiffres avec zéros à gauche. Positions absolues 27-33 dans LINCDE.",
     "0000012"),

    (31, "Libellé article (toujours espaces)", "LINCDE-LIBELL", "CHAR", 35,
     "Zone libellé article dans LINCDE. NON ALIMENTÉE INTENTIONNELLEMENT : "
     "MOVE SPACES TO ENR-LINCDE-LIBELL est exécuté systématiquement dans FORMAT-ORDERS-CDL "
     "(source ligne 1837), même si D.ART.LIBELL est disponible via TABLE-CDL (TCDL-LIBELL). "
     "Résultat : toujours 35 espaces dans le fichier de sortie. "
     "Règle à reproduire telle quelle en Java sans tenter de 'corriger' cet effacement. "
     "Positions absolues 35-69 dans LINCDE.",
     "(35 espaces)"),

    (32, "Code laboratoire interne (CODLABLAB)", "LINCDE-CODLABLAB", "CHAR", 4,
     "Code laboratoire interne au laboratoire, distinct de CODLAB (code CSP). "
     "Source : TCDL-CODLABLAB(IND-CDL) = D.CDL.CODLAB lu par CURCDL. "
     "MOVE TCDL-CODLABLAB(IND-CDL) TO ENR-LINCDE-CODLABLAB. "
     "Identifie le sous-laboratoire producteur de l'article. "
     "Exemple : '2951' déclenche la règle de remplacement de dépôt via D.PAR (paragraphe LECTURE-DEPLAB). "
     "Positions absolues 71-74 dans LINCDE.",
     "2951"),

    (33, "Rembourrage fin LINCDE", "FILLER", "CHAR", 38,
     "Zone de rembourrage de fin d'enregistrement LINCDE. Toujours espaces (INITIALIZE). "
     "Déclaré ENR-LINCDE-FILLER PIC X(38) dans le source COBOL (ligne 128-129). "
     "Positions absolues 75-112 dans LINCDE. "
     "Les positions 113-197 (85 chars) sont non adressées dans cette variante REDEFINES et restent espaces.",
     "(38 espaces)"),

    (34, "Somme quantités livrées du groupe", "FINCDE-SUMQTE", "BIGINT", 8,
     "Somme cumulée des quantités livrées (QTLCDE) pour toutes les lignes LINCDE du groupe laboratoire, par dépôt. "
     "W-SUMQTE-CO pour CO, W-SUMQTE-MO pour MO. "
     "Calculé dans FORMAT-ORDERS-CDL par ADD TCDL-QTELIV(IND-CDL) TO W-SUMQTE-CO (ou MO). "
     "Réinitialisé à 0 en début de chaque groupe laboratoire "
     "(STOCK-LIG-DESADV : MOVE 0 TO W-SUMQTE-CO W-SUMQTE-MO). "
     "MOVE W-SUMQTE-CO TO ENR-FINCDE-SUMQTE (ou MO). "
     "PIC 9(08), 8 chiffres. Positions absolues 8-15 dans FINCDE.",
     "00001440"),

    (35, "Nombre de lignes LINCDE du groupe", "FINCDE-NBLIG", "BIGINT", 8,
     "Nombre d'enregistrements LINCDE écrits pour ce groupe laboratoire et ce dépôt. "
     "W-NB-CDL-CO pour CO, W-NB-CDL-MO pour MO. "
     "Incrémenté dans STOCK-LIG-DESADV par ADD 1 TO W-NB-CDL-CO (ou MO) pour chaque entrée CODENR='3'. "
     "Réinitialisé à 0 à chaque changement de PREV-CODLAB. "
     "FINCDE est routé par PREV-CODDEP (dépôt du DERNIER enregistrement traité, pas le courant). "
     "PIC 9(08), 8 chiffres. Positions absolues 17-24 dans FINCDE.",
     "00000010"),

    (36, "Rembourrage fin FINCDE", "FILLER", "CHAR", 97,
     "Zone de rembourrage de fin d'enregistrement FINCDE. Toujours espaces (INITIALIZE). "
     "Déclaré PIC X(97) dans le source COBOL. "
     "Positions absolues 25-121 dans FINCDE. Positions 122-197 non adressées = espaces.",
     "(97 espaces)"),

    (37, "Nombre de commandes du fichier", "FINMES-NBCDE", "BIGINT", 8,
     "Nombre total de commandes ayant généré au moins un enregistrement LINCDE sur ce dépôt. "
     "NB-CDE-CO pour TRANSCO, NB-CDE-MO pour TRANSMO. "
     "Incrémenté dans STOCK-FIC-DESADV : IF W-NB-CDL-CO > 0 THEN ADD 1 TO NB-CDE-CO. "
     "W-NB-CDL-CO est lui-même le compteur de lignes CDL brutes (TRT-TROUVE-CDL), "
     "distinct de W-NB-CDL-CO réinitialisé dans STOCK-LIG-DESADV (compteur d'écriture). "
     "FINMES écrit une seule fois en pied de chaque fichier. "
     "PIC 9(08), 8 chiffres. Positions absolues 8-15 dans FINMES.",
     "00000025"),

    (38, "Rembourrage fin FINMES", "FILLER", "CHAR", 97,
     "Zone de rembourrage de fin d'enregistrement FINMES. Toujours espaces (INITIALIZE). "
     "Déclaré PIC X(97) dans le source COBOL. "
     "Positions absolues 16-112 dans FINMES. Positions 113-197 non adressées = espaces.",
     "(97 espaces)"),
]

# ═════════════════════════════════════════════════════════════════════════════
# TABLE 2 — Table cible S.CDE_FAC dans BD_CRM  (colonnes I-O, indices 9-15)
# 86 colonnes — données provenant du batch D02 ou calculées à partir de celui-ci.
# ═════════════════════════════════════════════════════════════════════════════
T2 = [
    (1, "Code laboratoire", "CODLAB", "CHAR", 4,
     "Code du laboratoire traité. Correspond au paramètre P-CODLAB transmis au batch "
     "et au filtre WHERE CODLAB = :W-CODLAB du curseur CURCDE/CURCDE_R. "
     "Clé principale de sélection des commandes dans BD_DEPOT.", "1234"),

    (2, "Code dépôt de livraison", "CODDEP", "CHAR", 2,
     "Code du dépôt de livraison : 'CO' = Cournon, 'MO' = Montélimar. "
     "Détermine le fichier cible (RMS_TRANSCO ou RMS_TRANSMO). "
     "Issu de CDL.CODDEP via TABLE-CDL (TCDL-CODDEP). "
     "Peut être remplacé pour le labo 2951 par FONCT(1:2) issu de D.PAR (paragraphe LECTURE-DEPLAB).", "CO"),

    (3, "Code sous-laboratoire", "SSLABO", "CHAR", 4,
     "Code laboratoire interne au laboratoire (CODLABLAB). "
     "Issu de CDL.CODLAB via TCDL-CODLABLAB. "
     "Identifie le sous-laboratoire producteur de chaque ligne de commande. "
     "Exemple : '2951' déclenche la règle de substitution de dépôt via D.PAR.", "2951"),

    (4, "Numéro de commande CSP", "NUMCDE", "BIGINT", None,
     "Numéro interne CSP de la commande. Issu de CDE.NUMCDE, lu via CURCDE/CURCDE_R. "
     "Forme avec NUMRAL la clé complète de la commande dans BD_DEPOT. "
     "Utilisé comme paramètre de liaison dans les curseurs CDL, MES et dans RECH-NUMDOC.", "1234567"),

    (5, "Numéro de reste-à-livrer", "NUMRAL", "BIGINT", None,
     "Numéro de reste-à-livrer CSP. Issu de CDE.NUMRAL. "
     "Avec NUMCDE, identifie de manière unique une commande dans CDE et ses lignes dans CDL. "
     "Présent dans le fichier MAJ (ENREG-NUMRAL).", "1"),

    (6, "Moyen de règlement", "MOYRGL", "CHAR", 2,
     "Code du moyen de règlement de la commande. Issu de CDE.MOYRGL. "
     "Alimenté dans FIC-TYPE1 du fichier MAJ, non présent dans TRANSCO/TRANSMO.", "VR"),

    (7, "Délai de règlement", "DELRGL", "CHAR", 3,
     "Code du délai de règlement accordé au client. Issu de CDE.DELRGL. "
     "Alimenté dans le fichier MAJ. Non présent dans TRANSCO/TRANSMO.", "030"),

    (8, "Code représentant commercial", "CODREP", "CHAR", 8,
     "Code du représentant commercial affecté à la commande. "
     "Issu de CDE.CODREP, indicateur NULL ICODREP dans CURCDE/CURCDE_R. "
     "Copié dans REFCDE-CODREP de TRANSCO/TRANSMO et dans FIC-CODREP du fichier MAJ "
     "(avec règle : SPACES si NUMCDE = FIC-CDELAB(1:7)).", "REP00123"),

    (9, "Date de commande", "DATCDE", "DATE VMS", None,
     "Date de création de la commande au format binaire VMS (PIC S9(11)V9(7) COMP). "
     "Issu de CDE.DATCDE via CURCDE/CURCDE_R. "
     "Converti en YYYYMMDD par SYS$ASCTIM+CVT-MOIS, "
     "puis tronqué à YYMMDD (FIC-DATCDE(3:6)) pour REFCDE-DATCDE dans TRANSCO/TRANSMO.",
     "15-AUG-2026 09:23:41.12"),

    (10, "Date du bon de livraison", "DATEBL", "DATE VMS", None,
     "Date du bon de livraison, format binaire VMS. "
     "Issu de CDE.DATEBL, indicateur NULL IDATEBL. "
     "Usage double : (1) critère de filtre dans CURCDE_R "
     "(DATEBL > :W-DATDEB AND DATEBL < :W-DATFIN) ; "
     "(2) si IDATEBL=-1 : FIC-DATEBL = PARAM-DATE-TXT (date du jour YYYYMMDD). "
     "Converti en YYYYMMDD pour le fichier MAJ.", "15-AUG-2026 09:23:41.12"),

    (11, "Taux d'escompte", "ESCOMP", "DECIMAL(6,4)", None,
     "Taux d'escompte accordé au client. Issu de CDE.ESCOMP. "
     "Non présent dans TRANSCO/TRANSMO. Alimenté dans le fichier MAJ.", "2.5000"),

    (12, "Frais de gestion", "FRGEST1", "DECIMAL(8,2)", None,
     "Frais de gestion appliqués à la commande. "
     "Non présent dans TRANSCO/TRANSMO. Alimenté dans le fichier MAJ.", "1500.00"),

    (13, "Statut de la commande", "FLAGCDE", "CHAR", 1,
     "Statut de la commande. Le curseur CURCDE filtre sur STATUT = 'CRV' (commande à valider). "
     "Seules les commandes au statut CRV sont traitées par D02.", "C"),

    (14, "Indicateur livrée", "FLAGLIV", "CHAR", 1,
     "Indicateur positionnant la commande comme livrée. Mis à jour après traitement par D02.", "O"),

    (15, "Indicateur facturée", "FLAGFAC", "CHAR", 1,
     "Indicateur de facturation. "
     "Non alimenté par D02, mis à jour par programme de facturation aval.", "N"),

    (16, "Indicateur bloquée", "FLAGBLO", "CHAR", 1,
     "Indicateur de blocage de la commande. Non alimenté par D02.", "N"),

    (17, "Indicateur supprimée", "FLAGSUP", "CHAR", 1,
     "Indicateur de suppression logique de la commande. Non alimenté par D02.", "N"),

    (18, "Motif de suppression", "MOTIFSUP", "CHAR", 3,
     "Code motif en cas de suppression logique. Non alimenté par D02.", "(vide)"),

    (19, "Indicateur différée", "FLAGDIF", "CHAR", 1,
     "Indicateur de commande différée (traitement reporté). Non alimenté par D02.", "N"),

    (20, "Indicateur reste-à-livrer", "FLAGRAL", "CHAR", 1,
     "Indicateur de commande partiellement livrée. Non alimenté directement par D02.", "O"),

    (21, "Indicateur ligne standard", "FLAGSTD", "CHAR", 1,
     "Indicateur de ligne standard. "
     "Lié à CDE.NBSTD, utilisé dans ADD NBSTD NBDTL TO W-NBCOL pour le fichier MAJ (FIC-NBCOL).", "S"),

    (22, "Indicateur ligne détail", "FLAGDET", "CHAR", 1,
     "Indicateur de ligne de détail. Lié à CDE.NBDTL. Voir FIC-NBCOL dans fichier MAJ.", "D"),

    (23, "Indicateur client générique livré", "GENCLI", "CHAR", 1,
     "Indicateur signalant un client générique côté livraison. "
     "Issu de CDE.GENCLILAB. Non alimenté dans TRANSCO/TRANSMO.", "N"),

    (24, "Libellé client générique livré", "LIBGENCLI", "CHAR", 40,
     "Libellé du client générique livré. Non alimenté par D02.", "(vide)"),

    (25, "Code client labo", "CLILAB", "CHAR", 10,
     "Code client côté laboratoire. Issu de CDE.CLILAB, mémorisé dans W-CLILAB dans TRT-TROUVE. "
     "Non présent dans TRANSCO/TRANSMO.", "CLILAB0001"),

    (26, "Indicateur client occasionnel", "FLAGOCC", "CHAR", 1,
     "Indicateur de client occasionnel. Non alimenté par D02.", "N"),

    (27, "Code client CSP livré", "CLICSP", "BIGINT", None,
     "Code client CSP du destinataire de la livraison. Issu de CDE.CLICSP ou CLI.CLICSP. "
     "Utilisé dans REFCDE (deux zones) : REFCDE-CLICSP (valeur brute 6 chiffres) "
     "et REFCDE-NOMLIV (astuce décalage sur 3 chiffres).", "1234"),

    (28, "Nom du client livré", "NOMLIV", "CHAR", 40,
     "Nom du destinataire de la livraison. Issu de CDO.NOMLIV (jointure CDE-CDO). "
     "Alimenté dans FIC-NOMLIV du fichier MAJ. Non présent dans TRANSCO/TRANSMO.", "PHARMACIE DUPONT"),

    (29, "Raison sociale client livré", "RAISOCL", "CHAR", 40,
     "Raison sociale du client livré. Issu de CDO.RAISOCL via lecture-cdo. "
     "Alimenté dans FIC-RAISOCL du fichier MAJ.", "PHARMACIE DUPONT SARL"),

    (30, "Adresse 1 client livré", "ADR1L", "CHAR", 40,
     "Première ligne d'adresse du client livré. Issu de CDO.ADR1L. Fichier MAJ uniquement.", "12 RUE DE LA PAIX"),

    (31, "Adresse 2 client livré", "ADR2L", "CHAR", 40,
     "Deuxième ligne d'adresse du client livré. Issu de CDO.ADR2L. Fichier MAJ uniquement.", "BP 123"),

    (32, "Ville client livré", "VILLEL", "CHAR", 32,
     "Ville du client livré. Issu de CDO.VILLEL. Alimenté dans FIC-VILLEL du fichier MAJ.", "PARIS"),

    (33, "Département client livré", "DEPARL", "CHAR", 2,
     "Code département du client livré. Non lu directement dans CDO par D02.", "75"),

    (34, "Code postal client livré", "CPOSTL", "CHAR", 5,
     "Code postal du client livré. Issu de CDO.CPOSTL. Alimenté dans FIC-CPOSTL du fichier MAJ.", "75001"),

    (35, "Code UGA", "UGA", "CHAR", 3,
     "Code de l'Unité de Gestion Administrative. Non alimenté par D02.", "(vide)"),

    (36, "Code UGA variante 746", "UGA746", "CHAR", 5,
     "Code UGA pour la variante 746. Non alimenté par D02.", "(vide)"),

    (37, "Code client CSP facturé", "CLIFAC", "BIGINT", None,
     "Code client CSP côté facturation. Non alimenté directement par D02.", "(vide)"),

    (38, "Nom du client facturé", "NOMFAC", "CHAR", 40,
     "Nom du client facturé. Non alimenté par D02.", "(vide)"),

    (39, "Raison sociale client facturé", "RAISOCF", "CHAR", 40,
     "Raison sociale du client facturé. Non alimenté par D02.", "(vide)"),

    (40, "Adresse 1 client facturé", "ADR1F", "CHAR", 40,
     "Première ligne d'adresse du client facturé. Non alimenté par D02.", "(vide)"),

    (41, "Adresse 2 client facturé", "ADR2F", "CHAR", 40,
     "Deuxième ligne d'adresse. Non alimenté par D02.", "(vide)"),

    (42, "Ville client facturé", "VILLEF", "CHAR", 32,
     "Ville du client facturé. Non alimenté par D02.", "(vide)"),

    (43, "Département client facturé", "DEPARF", "CHAR", 2,
     "Code département du client facturé. Non alimenté par D02.", "(vide)"),

    (44, "Code postal client facturé", "CPOSTF", "CHAR", 5,
     "Code postal du client facturé. Non alimenté par D02.", "(vide)"),

    (45, "Code client CSP payeur", "CLIPAY", "BIGINT", None,
     "Code client CSP côté règlement. Non alimenté par D02.", "(vide)"),

    (46, "Nom du client payeur", "NOMPAY", "CHAR", 40,
     "Nom du client payeur. Non alimenté par D02.", "(vide)"),

    (47, "Raison sociale client payeur", "RAISPAY", "CHAR", 40,
     "Raison sociale du client payeur. Non alimenté par D02.", "(vide)"),

    (48, "Code client CSP groupe", "CLIGRP", "BIGINT", None,
     "Code client CSP du groupe auquel appartient le client. Non alimenté par D02.", "(vide)"),

    (49, "Nom du client groupe", "NOMGRP", "CHAR", 40,
     "Nom du client groupe. Non alimenté par D02.", "(vide)"),

    (50, "Raison sociale client groupe", "RAISGRP", "CHAR", 40,
     "Raison sociale du client groupe. Non alimenté par D02.", "(vide)"),

    (51, "Type de facture", "TYPNFA", "CHAR", 2,
     "Type de facturation. Issu de CDE.TYPNFA, indicateur ITYPNFA. "
     "Contribue à la composition de FIC-NUMFAC dans le fichier MAJ "
     "(MOVE TYPNFA TO FIC-NUMFAC(1:2)).", "FA"),

    (52, "Indicateur quantième", "QUANTA", "CHAR", 1,
     "Dernier chiffre de l'année de facturation. Issu de CDE.QUANTA, indicateur IQUANTA. "
     "MOVE QUANTA TO FIC-NUMFAC(3:1) dans le fichier MAJ.", "6"),

    (53, "Mois de facturation", "MOISFA", "CHAR", 2,
     "Mois de facturation. Issu de CDE.MOISFA, indicateur IMOISFA. "
     "MOVE MOISFA TO FIC-NUMFAC(4:2) dans le fichier MAJ.", "08"),

    (54, "Compteur de facturation", "CPTFAC", "BIGINT", None,
     "Numéro séquentiel de facture. Issu de CDE.CPTFAC, indicateur ICPTFAC. "
     "MOVE W-CPTFAC TO FIC-NUMFAC(6:5) dans le fichier MAJ. "
     "INSPECT FIC-NUMFAC REPLACING ALL SPACE BY ZERO.", "12345"),

    (55, "Date de facturation", "DATFAC", "DATE VMS", None,
     "Date de facturation. Non alimenté directement par D02.", "(vide)"),

    (56, "Code représentant national", "CODREPN", "CHAR", 8,
     "Code représentant national. Non alimenté par D02.", "(vide)"),

    (57, "Code représentant régional", "CODREPR", "CHAR", 8,
     "Code représentant régional. Non alimenté par D02.", "(vide)"),

    (58, "Indicateur regroupement facturation", "REGFAC", "CHAR", 1,
     "Indicateur de regroupement sur la facture. Non alimenté par D02.", "(vide)"),

    (59, "Nom du représentant", "NOMREP", "CHAR", 40,
     "Nom du représentant commercial. Non alimenté par D02.", "(vide)"),

    (60, "Nom du représentant national", "NOMREPN", "CHAR", 40,
     "Nom du représentant national. Non alimenté par D02.", "(vide)"),

    (61, "Nom du représentant régional", "NOMREPR", "CHAR", 40,
     "Nom du représentant régional. Non alimenté par D02.", "(vide)"),

    (62, "Indicateur en-cours", "FLAGENCOURS", "CHAR", 1,
     "Indicateur de commande en cours de traitement. Non alimenté par D02.", "(vide)"),

    (63, "Quantité en-cours", "QTENCOURS", "BIGINT", None,
     "Quantité en cours de préparation. Non alimenté par D02.", "(vide)"),

    (64, "Statut de l'en-cours", "STATENCOURS", "CHAR", 3,
     "Statut de la ligne en cours. Non alimenté par D02.", "(vide)"),

    (65, "Taux de remise", "REMISEE", "DECIMAL(10,4)", None,
     "Taux de remise appliqué à la commande. Non alimenté par D02.", "(vide)"),

    (66, "Montant hors taxes", "MTHT", "DECIMAL(10,2)", None,
     "Montant total hors taxes de la commande. Non alimenté par D02.", "(vide)"),

    (67, "Date de saisie", "DATSAI", "DATE VMS", None,
     "Date de saisie de la commande. Issu de CDE.DATSAI, converti en YYYYMMDD "
     "par SYS$ASCTIM+CVT-MOIS pour alimenter FIC-DATSAI dans le fichier MAJ.",
     "15-AUG-2026 09:23:41.12"),

    (68, "Date d'échéance de règlement", "DATECH", "DATE VMS", None,
     "Date d'échéance de règlement. Non alimenté directement par D02.", "(vide)"),

    (69, "Code CIP point de vente", "CIPPDV", "BIGINT", None,
     "Code CIP du point de vente. Issu de CLI.CIPPDV avec indicateur NULL ICIPPDV dans get-cli. "
     "Alimenté dans le fichier MAJ.", "1234567"),

    (70, "Indicateur client générique facturé", "GENCLIFAC", "CHAR", 1,
     "Indicateur client générique côté facturation. Non alimenté par D02.", "(vide)"),

    (71, "Libellé client générique facturé", "LIBGENCLIFAC", "CHAR", 40,
     "Libellé client générique facturé. Non alimenté par D02.", "(vide)"),

    (72, "Code représentant court", "CODREPC", "CHAR", 5,
     "Code représentant abrégé. Non alimenté par D02.", "(vide)"),

    (73, "Nom du représentant court", "NOMREPC", "CHAR", 80,
     "Nom du représentant abrégé. Non alimenté par D02.", "(vide)"),

    (74, "Date de livraison souhaitée", "DTLIVS", "DATE VMS", None,
     "Date de livraison souhaitée par le client. Non alimenté directement par D02.", "(vide)"),

    (75, "Poids total de la commande", "POIDS", "DECIMAL(11,4)", None,
     "Poids total de la commande. Calculé dans STOCK-FIC-DESADV : "
     "ADD PDSSTD IN CDE PDSDTL IN CDE TO W-PDS → MOVE W-PDS TO FIC-PDS dans le fichier MAJ.",
     "12345.6789"),

    (76, "Numéro commande laboratoire", "CDELAB", "CHAR", 22,
     "Numéro de commande côté laboratoire. Issu de CDE.CDELAB. "
     "Alimenté dans REFCDE-CDELAB (TRANSCO/TRANSMO) et ENREG-CDELAB (fichier MAJ).",
     "LAB2024-00123456789   "),

    (77, "Référence interne commande", "REFCDE", "CHAR", 35,
     "Référence interne CSP. Issu de CDE.REFCDE (:w-refcde). "
     "Alimenté dans REFCDE-REFCDE (TRANSCO/TRANSMO).",
     "CSP-REF-2024-001234567890123456789"),

    (78, "Dossier d'expédition", "DOSEXP", "CHAR", 10,
     "Numéro du dossier d'expédition. "
     "Calculé par appel externe D00_NUMEXP (CALL D00_NUMEXP USING PARAM-BASNUMCDE → WP-NUMEXP). "
     "Non présent dans TRANSCO/TRANSMO.", "EXP0012345"),

    (79, "Volume standard", "VOLSTD", "BIGINT", None,
     "Volume des colis standard. Issu de CDE.VOLSTD. ADD VOLSTD VOLDTL → FIC-VOL dans fichier MAJ.", "1200"),

    (80, "Volume détail", "VOLDTL", "BIGINT", None,
     "Volume des colis de détail. Issu de CDE.VOLDTL. Additionné à VOLSTD pour FIC-VOL.", "300"),

    (81, "Libellé pays", "LIBPAYS", "CHAR", 80,
     "Libellé du pays du client. Issu de CLI.PAYS via get-cli / get-cli-pay. "
     "Alimenté dans le fichier MAJ.", "FRANCE"),

    (82, "Code promotion", "PROMOS", "CHAR", 4,
     "Code de l'opération promotionnelle. Non alimenté par D02.", "(vide)"),

    (83, "Nombre de colis", "NBCOLIS", "BIGINT", None,
     "Nombre de colis. ADD NBSTD IN CDE NBDTL IN CDE TO W-NBCOL → MOVE W-NBCOL TO FIC-NBCOL dans MAJ.", "24"),

    (84, "Nombre de palettes", "NBPAL", "BIGINT", None,
     "Nombre de palettes. Non alimenté directement par D02.", "(vide)"),

    (85, "Date de livraison réelle", "DTLIVR", "DATE VMS", None,
     "Date effective de livraison. Non alimenté par D02 (mis à jour par programme aval).", "(vide)"),

    (86, "Indicateur expédiée", "FLAGEXP", "CHAR", 1,
     "Indicateur de commande expédiée. "
     "Positionné par le programme aval après traitement de D02.", "O"),
]

# ═════════════════════════════════════════════════════════════════════════════
# TABLE 3 — Tables sources BD_DEPOT  (colonnes Q-W, indices 17-23)
# 28 colonnes sources des tables CDE, CDL, MES, ART.
# ═════════════════════════════════════════════════════════════════════════════
T3 = [
    (1, "Date commande (source)", "DATCDE", "DATE VMS", None,
     "Table CDE. Date de création de la commande, format binaire VMS (PIC S9(11)V9(7) COMP). "
     "Lu via CURCDE ou CURCDE_R. "
     "Transformé en YYYYMMDD par CALL SYS$ASCTIM + CVT-MOIS + INSPECT, "
     "puis tronqué à YYMMDD via FIC-DATCDE(3:6) pour alimenter REFCDE-DATCDE dans TRANSCO/TRANSMO.",
     "15-AUG-2026 09:23:41.12"),

    (2, "Date bon de livraison (source)", "DATEBL", "DATE VMS", None,
     "Table CDE. Format binaire VMS. Indicateur NULL IDATEBL dans CURCDE. "
     "Rôle double : (1) dans CURCDE_R, critère de filtre avec bornes strictement exclues "
     "(DATEBL > :W-DATDEB AND DATEBL < :W-DATFIN) ; "
     "(2) si IDATEBL = -1 : FIC-DATEBL = PARAM-DATE-TXT (date du jour YYYYMMDD). "
     "Converti en YYYYMMDD pour le fichier MAJ.", "15-AUG-2026 09:23:41.12"),

    (3, "Date réception commande (source)", "DATREC", "DATE VMS", None,
     "Table CDE. Date et heure de réception de la commande par CSP. "
     "Dans CURCDE : lu AVEC indicateur NULL (SQLSTATE '22002' traité comme succès — comportement B05). "
     "Dans CURCDE_R : lu SANS indicateur NULL. "
     "Converti par SYS$ASCTIM → W-DATE-ASCII 23 chars, "
     "stocké dans FIC-DATREC, copié dans REFCDE-DATREC de TRANSCO/TRANSMO.",
     "15-AUG-2026 09:23:41.12"),

    (4, "Date édition bon préparation (source)", "DATEBP", "DATE VMS", None,
     "Table CDE. Format binaire VMS. Indicateur NULL IDATEBP. "
     "Si IDATEBP = -1 : FIC-DATEBP = PARAM-DATE-TXT (date du jour YYYYMMDD). "
     "Sinon : converti en YYYYMMDD par SYS$ASCTIM + CVT-MOIS. "
     "Copié dans REFCDE-DATEBP de TRANSCO/TRANSMO.", "15-AUG-2026 09:23:41.12"),

    (5, "Référence labo commande (source)", "CDELAB", "CHAR", 22,
     "Table CDE. Numéro de commande attribué par le laboratoire. "
     "Lu via CURCDE/CURCDE_R, copié dans REFCDE-CDELAB (22 chars) de TRANSCO/TRANSMO "
     "et dans ENREG-CDELAB (10 chars) du fichier MAJ.", "LAB2024-00123456789   "),

    (6, "Référence interne commande (source)", "REFCDE", "CHAR", 35,
     "Table CDE. Référence interne CSP. Lu via CURCDE/CURCDE_R "
     "dans la variable hôte :w-refcde (PIC X(35), déclarée ligne 367 du source COBOL). "
     "Copié dans REFCDE-REFCDE (35 chars) de TRANSCO/TRANSMO.",
     "CSP-REF-2024-001234567890123456789"),

    (7, "Type de commande (source)", "TYPCDE", "CHAR", 2,
     "Table CDE. Code du type de commande. Lu via CURCDE/CURCDE_R. "
     "MOVE TYPCDE IN CDE TO ENR-REFCDE-TYPCDE. "
     "Exemples : ST = standard, RA = reste-à-livrer, GR = gratuit. "
     "Positions absolues 74-75 de REFCDE dans TRANSCO/TRANSMO.", "ST"),

    (8, "Code client livré CSP (source)", "CLICSP", "BIGINT", None,
     "Table CDE. Code numérique du client destinataire (PIC 9(06)). "
     "Deux usages dans REFCDE : (1) REFCDE-CLICSP : valeur brute 6 chiffres avec zéros à gauche ; "
     "(2) REFCDE-NOMLIV : astuce décalage — MOVE CLICSP TO NOMLIV "
     "puis MOVE NOMLIV(4:6) TO NOMLIV supprime les 3 zéros de tête ('001234' → '234' + 32 espaces).",
     "001234"),

    (9, "Code opérateur saisie commande (source)", "CODOPE", "CHAR", 3,
     "Table CDE. Identifiant de l'opérateur ou du système ayant créé la commande. "
     "Lu via CURCDE/CURCDE_R. MOVE CODOPE IN CDE TO ENR-REFCDE-CODOPE. "
     "Positions absolues 120-122 de REFCDE dans TRANSCO/TRANSMO.", "OPE"),

    (10, "Indicateur saisie manuelle (source)", "CDESAISIE", "CHAR", 1,
     "Table CDE. Indicateur O/N signalant si la commande a été saisie manuellement. "
     "Lu via CURCDE/CURCDE_R. MOVE CDESAISIE IN CDE TO ENR-REFCDE-CDESAISIE. "
     "Position absolue 135 de REFCDE dans TRANSCO/TRANSMO.", "O"),

    (11, "Mode trafic transport (source)", "TRAFIC", "CHAR", 1,
     "Table CDE. Mode de transport : E = express, N = normal. "
     "Lu via CURCDE/CURCDE_R. "
     "Règle de substitution : IF TRAFIC IN CDE = SPACES MOVE 'N' TO ENR-REFCDE-TRAFIC. "
     "Position absolue 161 de REFCDE dans TRANSCO/TRANSMO.", "N"),

    (12, "Code opérateur web (source)", "CODSAI", "CHAR", 7,
     "Table CDE. Identifiant de l'interface web ou de l'opérateur en ligne ayant passé la commande. "
     "Lu via CURCDE/CURCDE_R. MOVE CODSAI IN CDE TO ENR-REFCDE-CODSAI. "
     "Positions absolues 172-178 de REFCDE dans TRANSCO/TRANSMO.", "WEBPDA1"),

    (13, "Code sous-traitant préparation (source)", "CODSTR", "CHAR", 4,
     "Table CDE. Code du prestataire chargé de la préparation des colis. "
     "Lu via CURCDE/CURCDE_R, indicateur NULL ICODSTR. "
     "MOVE CODSTR IN CDE TO ENR-REFCDE-CODSTR. "
     "Positions absolues 180-183 de REFCDE dans TRANSCO/TRANSMO.", "EXPL"),

    (14, "Code structure expédition (source)", "STREXP", "CHAR", 4,
     "Table CDE. Code de la structure d'expédition (transporteur sous-traitant final). "
     "Lu via CURCDE/CURCDE_R, indicateur NULL ISTREXP. "
     "MOVE STREXP IN CDE TO ENR-REFCDE-STREXP. "
     "Positions absolues 185-188 de REFCDE dans TRANSCO/TRANSMO.", "CHRO"),

    (15, "Code représentant (source)", "CODREP", "CHAR", 8,
     "Table CDE. Code représentant commercial. "
     "Lu via CURCDE/CURCDE_R, indicateur NULL ICODREP. "
     "MOVE CODREP IN CDE TO ENR-REFCDE-CODREP. Positions absolues 190-197 de REFCDE. "
     "Règle MAJ : FIC-CODREP = SPACES si NUMCDE = FIC-CDELAB(1:7) ; sinon W-CODREP.", "REP00123"),

    (16, "Code laboratoire interne (source)", "CODLABLAB", "CHAR", 4,
     "Table CDE. Code laboratoire interne distinct de CODLAB CSP. "
     "Lu via CURCDE/CURCDE_R, indicateur NULL ICODLABLAB. "
     "Propagé dans TABLE-CDL (TCDL-CODLABLAB) puis copié dans LINCDE-CODLABLAB (4 chars) de TRANSCO/TRANSMO. "
     "La valeur '2951' déclenche le remplacement de CODDEP via D.PAR (LECTURE-DEPLAB).", "2951"),

    (17, "Texte commentaire BL/RT (source)", "COMMENT", "CHAR", 35,
     "Table MES. Libellé du message attaché à la commande. "
     "Lu par RECHERCHE-MES-BL-RT : "
     "SELECT COMMENT FROM D.MES WHERE CODMES=CDE.CODMES AND TYPDOC='BL' ; "
     "si absent, SELECT ... AND TYPDOC='RT'. "
     "Stocké dans W-COMMENT (PIC X(35)), copié dans LINTXT-MESSAGE (114 chars, padding espaces). "
     "Si aucun message : 114 espaces. Un enregistrement LINTXT est TOUJOURS écrit.",
     "Livraison avant 12h - quai B       "),

    (18, "Type de document message (source)", "TYPDOC", "CHAR", 3,
     "Table MES. Type de document du message : 'BL ' (bon de livraison) ou 'RT ' (retour). "
     "Critère de sélection prioritaire dans RECHERCHE-MES-BL-RT. "
     "Détermine aussi TXTCDE-TYPDOC : article QU00013 → 'RT', article QU000132 → 'BL'.", "BL "),

    (19, "Code article CSP (source)", "CODART", "CHAR", 10,
     "Table CDL. Code article CSP de la ligne de commande. "
     "Lu via CURCDL (SELECT L.NUMCDE, L.NUMRAL, CODART... FROM D.CDL JOIN D.ART). "
     "Stocké dans TCDL-CODART dans TABLE-CDL. "
     "Copié dans LINCDE-CODART (10 chars) de TRANSCO/TRANSMO pour les entrées CODENR='3'. "
     "Articles spéciaux : QU00013 → TXTCDE/RT, QU000132 → TXTCDE/BL.", "3400936750"),

    (20, "Libellé produit / texte spécial (source)", "LIBELL", "CHAR", 35,
     "Table CDL. Libellé de la ligne de commande (PIC X(35) via domaine LIBEL50 tronqué). "
     "Rôle dual selon CODART : "
     "(1) Article ordinaire : lu mais NON copié dans TRANSCO/TRANSMO "
     "(LINCDE-LIBELL = toujours SPACES) ; "
     "(2) Article QU00013 : copié dans W-LIB-HORAIRE (80 chars), alimente TXTCDE-MESSAGE/RT ; "
     "(3) Article QU000132 : copié dans W-LIB-CONTACT (80 chars), alimente TXTCDE-MESSAGE/BL.",
     "Livraison 8h-12h lundi-vendredi    "),

    (21, "Quantité livrée par ligne (source)", "QTLCDE", "BIGINT", None,
     "Table CDL. Quantité de la ligne effectivement livrée. "
     "Cumulée par groupe (NUMLIG, CODART) dans TCDL-QTELIV via ADD CDL.QTLCDE TO TCDL-QTELIV. "
     "Copié dans ENR-LINCDE-QTCCDE (7 chiffres) malgré le nom trompeur 'QTCCDE' de la zone de sortie. "
     "Règle critique B — comportement délibéré à reproduire tel quel en Java sans 'corriger'.", "144"),

    (22, "Quantité gratuite par ligne (source)", "QTCGRT", "BIGINT", None,
     "Table CDL. Quantité gratuite accordée sur la ligne. "
     "Cumulée avec QTCECH dans TCDL-QTCGRT : ADD CDL.QTCGRT CDL.QTCECH TO TCDL-QTCGRT. "
     "Copié dans ENR-LINCDE-QTCGRT (7 chiffres) de TRANSCO/TRANSMO.", "12"),

    (23, "Quantité échangée (incluse dans gratuit)", "QTCECH", "BIGINT", None,
     "Table CDL. Quantité échangée (produits remplacés à titre de réclamation). "
     "Fusionnée avec QTCGRT dans le cumul TCDL-QTCGRT : ADD CDL.QTCGRT CDL.QTCECH TO TCDL-QTCGRT. "
     "Pas de champ dédié dans le fichier de sortie : noyée dans LINCDE-QTCGRT avec la quantité gratuite.", "2"),

    (24, "Numéro de lot fabricant (source)", "LOTFAB", "CHAR", 12,
     "Table CDL. Numéro de lot de fabrication (PIC X(12)). "
     "Copié dans TCDL-LOTFAB de TABLE-CDL. "
     "Règle de conservation (paragraphe TRT-CDL-CUMULEE / TRAITEMENT-SANS-LOT) : "
     "si ART.GESLOT IN {'4','5'} → LOTFAB conservé dans l'entrée CODENR='3' ; "
     "sinon → MOVE SPACES TO TCDL-LOTFAB. "
     "L'entrée CODENR='1' a toujours LOTFAB = SPACES. "
     "Non présent dans TRANSCO/TRANSMO ; présent dans fichier MAJ (FIC-LOTFAB).", "LOT2026-001 "),

    (25, "Numéro de ligne de commande (source)", "NUMLIG", "BIGINT", None,
     "Table CDL. Numéro de ligne dans la commande (PIC 9(4)). "
     "Clé de regroupement avec CODART dans TABLE-CDL : "
     "tant que (NUMLIG, CODART) est identique entre deux lignes CURCDL successives, "
     "les quantités sont cumulées dans la même entrée TABLE-CDL. "
     "Un changement de NUMLIG ou CODART génère une nouvelle paire d'entrées "
     "(CODENR='1' puis '3').", "10"),

    (26, "Code dépôt livraison (source)", "CODDEP", "CHAR", 2,
     "Table CDL. Code du dépôt de livraison de la ligne : "
     "'CO' = Cournon, 'MO' = Montélimar. "
     "Détermine le routage de chaque enregistrement LINCDE, REFCDE, LINTXT, TXTCDE "
     "vers TRANSCO ou TRANSMO. "
     "FINCDE est routé par PREV-CODDEP (dépôt du DERNIER enregistrement traité, pas le courant). "
     "Pour le laboratoire 2951 : CODDEP peut être remplacé par PAR.FONCT(1:2) "
     "(SELECT FONCT FROM D.PAR WHERE CODENT='LAB' AND CHAMPS='DEPLAB' AND ARGUM=WS-ARGUM).", "CO"),

    (27, "Code laboratoire ligne CDL (source)", "CODLAB", "CHAR", 4,
     "Table CDL. Code laboratoire de la ligne de commande. "
     "Tri primaire du curseur CURCDL (ORDER BY CODDEP, CODLAB). "
     "Stocké dans TCDL-CODLAB. "
     "Détecte les changements de groupe laboratoire dans STOCK-LIG-DESADV : "
     "IF TCDL-CODLAB(IND-CDL) NOT= PREV-CODLAB → écrire FINCDE (si pas première entrée) "
     "puis nouveau REFCDE + LINTXT + TXTCDE(s).", "1234"),

    (28, "Mode de gestion des lots article (source)", "GESLOT", "CHAR", 1,
     "Table ART. Mode de gestion des lots pour l'article (PIC X(1)). "
     "Lu via JOIN D.CDL D.ART ON CODART dans CURCDL. "
     "Règle de conservation du lot dans TABLE-CDL : "
     "valeurs '4' ou '5' → LOTFAB de CDL conservé dans TCDL-LOTFAB (entrée CODENR='3') ; "
     "toute autre valeur → TCDL-LOTFAB = SPACES (paragraphe TRAITEMENT-SANS-LOT). "
     "L'entrée CODENR='1' a toujours TCDL-LOTFAB = SPACES indépendamment de GESLOT.", "4"),
]

# ─────────────────────────────────────────────────────────────────────────────
# Effacement et réécriture
# ─────────────────────────────────────────────────────────────────────────────
clear_table(1)   # T1 — col A-G
clear_table(9)   # T2 — col I-O
clear_table(17)  # T3 — col Q-W

write_table(T1, start_col=1)
write_table(T2, start_col=9)
write_table(T3, start_col=17)

wb.save(XLSX)

# ─────────────────────────────────────────────────────────────────────────────
# Vérification
# ─────────────────────────────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(XLSX)
ws2 = wb2.active

def count_rows(sc):
    return sum(1 for r in range(3, 92) if ws2.cell(row=r, column=sc).value is not None)

t1n = count_rows(1)
t2n = count_rows(9)
t3n = count_rows(17)

print(f"OK — T1:{t1n} lignes | T2:{t2n} lignes | T3:{t3n} lignes")
assert t1n == 38, f"T1 attendu 38, obtenu {t1n}"
assert t2n == 86, f"T2 attendu 86, obtenu {t2n}"
assert t3n == 28, f"T3 attendu 28, obtenu {t3n}"
print("Assertions OK — fichier enregistré.")
