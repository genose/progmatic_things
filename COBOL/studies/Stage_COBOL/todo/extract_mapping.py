#!/usr/bin/env python3
"""
extract_mapping.py

Outil reutilisable pour generer un mapping de champs (data dictionary) a partir :
    - d'un copybook/programme COBOL (PIC clauses en WORKING-STORAGE ou LINKAGE)
  - d'extraits DDL Oracle Rdb (RMU/EXTRACT) au format
        create table TABLE (
            COLONNE DOMAINE
                comment is
                  'texte du commentaire',
            ...
        );

Il applique les conventions retenues sur ce projet :
  - PIC 9(n) / S9(n) sans V          -> Format=BIGINT
  - PIC 9(a)V9(b) / S9(a)V9(b)       -> Format=DECIMAL(a+b,b)
  - PIC X(23)                        -> Format=DATE VMS (heuristique : timestamp VMS 23 car.)
  - PIC X(n) (n != 23)               -> Format=CHAR, Taille=n

Usage:
    python3 extract_mapping.py --cob FILE.COB --ddl TAB_DEPOT.TXT [--ddl TAB_CRM.TXT ...] \
        --table CDE --table CDL --table MES --prefix ENR-REFCDE- \
        [--section working-storage|linkage|all] \
        [--xlsx Mapping.xlsx --sheet-col Q --sheet-start-row 3]

Sans --xlsx, le resultat est affiche en CSV sur stdout (N;Nom;Format;Taille;Commentaire;Exemple).
"""

import argparse
import csv
import re
import sys
import unicodedata

PIC_RE = re.compile(
    r"^\s*\d+\s+([\w-]+)(?:\s+REDEFINES\s+[\w-]+)?\s+PIC\s+(S?9[\d\(\)V]*|X\(\d+\)|X)\s*\.?\s*$",
    re.IGNORECASE,
)

ALPHA = "ABCDEFGHIJ0123456789"
DIGITS = "1234567890"


def strip_accents(text):
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(c for c in normalized if not unicodedata.combining(c))


def pic_to_format(pic):
    """Traduit une clause PIC en (Format, Taille) selon les conventions du projet."""
    pic = pic.upper().replace(" ", "")
    m = re.match(r"^S?9\((\d+)\)V9\((\d+)\)$", pic)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        return f"DECIMAL({a + b},{b})", None
    m = re.match(r"^S?9\((\d+)\)$", pic)
    if m:
        return "BIGINT", None
    m = re.match(r"^X\((\d+)\)$", pic)
    if m:
        n = int(m.group(1))
        if n == 23:
            return "DATE VMS", None
        return "CHAR", n
    if pic == "X":
        return "CHAR", 1
    return "UNKNOWN", None


def parse_cobol_fields(cob_path, prefix=None, section="working-storage"):
    """Extrait les champs PIC de WORKING-STORAGE, LINKAGE ou des deux sections."""
    fields = []
    with open(cob_path, encoding="latin-1") as f:
        lines = f.readlines()

    selected_sections = {
        "working-storage": {"WORKING-STORAGE"},
        "linkage": {"LINKAGE"},
        "all": {"WORKING-STORAGE", "LINKAGE"},
    }[section]
    current_section = None
    for line in lines:
        stripped = line.strip()
        upper = stripped.upper()
        section_match = re.match(r"^(WORKING-STORAGE|LINKAGE)\s+SECTION\s*\.", upper)
        if section_match:
            current_section = section_match.group(1)
            continue
        if upper.startswith("PROCEDURE DIVISION"):
            break
        if current_section not in selected_sections:
            continue
        m = PIC_RE.match(line.rstrip("\n"))
        if not m:
            continue
        name, pic = m.group(1), m.group(2)
        if name.upper() == "FILLER":
            continue
        if prefix and not name.upper().startswith(prefix.upper()):
            continue
        short_name = name[len(prefix):] if prefix else name
        short_name = short_name.lstrip("-")
        fmt, taille = pic_to_format(pic)
        fields.append({"nom": short_name, "pic": pic, "format": fmt, "taille": taille})
    return fields


COL_LINE_RE = re.compile(r'^\s*"?([A-Z_][A-Z0-9_]*)"?(?:\s+\S.*)?$')


def parse_ddl_tables(ddl_path, tables):
    """Extrait {TABLE: {COLONNE: commentaire}} pour les tables demandees d'un extrait DDL Rdb.

    Parcourt ligne a ligne le corps de chaque "create table X (...)" : une colonne demarre
    par une ligne "NOM [DOMAINE]" en MAJUSCULES, suivie optionnellement de "comment is" + texte
    entre quotes. Le "comment is ... ;" final (commentaire de table, hors parentheses) est ignore.
    """
    wanted = {t.upper() for t in tables}
    with open(ddl_path, encoding="latin-1") as f:
        text = f.read()

    starts = list(re.finditer(r"create\s+table\s+(\w+)\s*\(", text, re.IGNORECASE))
    result = {}
    for i, m in enumerate(starts):
        table_name = m.group(1).upper()
        if table_name not in wanted:
            continue
        body_end = starts[i + 1].start() if i + 1 < len(starts) else len(text)
        body = text[m.end():body_end]

        columns = {}
        current_col = None
        pending_comment = None
        expecting_text = False
        for line in body.splitlines():
            s = line.strip()
            if not s:
                continue
            col_match = COL_LINE_RE.match(line)
            if col_match and not s.lower().startswith("comment"):
                if current_col is not None:
                    columns.setdefault(current_col, pending_comment)
                current_col = col_match.group(1).upper()
                pending_comment = None
                expecting_text = False
                if s.rstrip(",").endswith(")"):
                    columns.setdefault(current_col, pending_comment)
                    current_col = None
                    break
                continue
            if s.lower().startswith("comment is"):
                if current_col is None:
                    break  # commentaire de table (hors parentheses), pas une colonne
                expecting_text = True
                continue
            if expecting_text:
                text_match = re.search(r"'([^']*)'", s)
                if text_match:
                    pending_comment = text_match.group(1).strip()
                    expecting_text = False
                    if re.search(r"'\s*\)", s):
                        columns.setdefault(current_col, pending_comment)
                        current_col = None
                        break
                continue
        if current_col is not None:
            columns.setdefault(current_col, pending_comment)
        result[table_name] = columns
    return result


def match_comment(short_name, ddl_by_table):
    """Cherche short_name dans les tables DDL (dans l'ordre donne) et retourne '(TABLE) commentaire'."""
    for table_name, columns in ddl_by_table.items():
        comment = columns.get(short_name.upper())
        if comment:
            clean = strip_accents(comment).lower().strip()
            return f"({table_name}) {clean}"
    return None


def example_for(fmt, taille):
    if not fmt or fmt == "UNKNOWN":
        return None
    if fmt == "CHAR":
        n = int(taille) if taille else 4
        return (ALPHA * (n // len(ALPHA) + 1))[:n]
    if fmt == "BIGINT":
        return "12345678"
    if fmt == "DATE VMS":
        return "15-AUG-2026 09:23:41.12"
    m = re.match(r"DECIMAL\((\d+),(\d+)\)", fmt)
    if m:
        p, s = int(m.group(1)), int(m.group(2))
        intdigits = p - s
        intpart = DIGITS[:intdigits] if intdigits > 0 else "0"
        decpart = DIGITS[:s] if s > 0 else ""
        return f"{intpart}.{decpart}" if s else intpart
    return None


def build_rows(cob_path, ddl_paths, tables, prefix, section="working-storage"):
    fields = parse_cobol_fields(cob_path, prefix, section)
    ddl_by_table = {}
    for ddl_path in ddl_paths:
        for table_name, columns in parse_ddl_tables(ddl_path, tables).items():
            ddl_by_table.setdefault(table_name, {}).update(columns)
    # respecter l'ordre de priorite donne par --table
    ordered_ddl = {t.upper(): ddl_by_table.get(t.upper(), {}) for t in tables}

    rows = []
    for i, field in enumerate(fields, start=1):
        comment = match_comment(field["nom"], ordered_ddl)
        example = example_for(field["format"], field["taille"])
        rows.append(
            {
                "n": i,
                "nom": field["nom"],
                "format": field["format"],
                "taille": field["taille"],
                "commentaire": comment,
                "exemple": example,
            }
        )
    return rows


def write_csv(rows, out=sys.stdout):
    writer = csv.writer(out, delimiter=";")
    writer.writerow(["N", "Nom", "Format", "Taille", "Commentaire", "Exemple"])
    for row in rows:
        writer.writerow([row["n"], row["nom"], row["format"], row["taille"], row["commentaire"], row["exemple"]])


def write_xlsx(xlsx_path, rows, start_col_letter, start_row, header_row):
    import openpyxl
    from openpyxl.utils import column_index_from_string

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    base_col = column_index_from_string(start_col_letter)

    headers = ["N°", "Nom", "Format", "Taille", "Commentaire", "Exemple"]
    for offset, header in enumerate(headers):
        ws.cell(row=header_row, column=base_col + offset).value = header

    for offset, row in enumerate(rows):
        r = start_row + offset
        ws.cell(row=r, column=base_col + 0).value = row["n"]
        ws.cell(row=r, column=base_col + 1).value = row["nom"]
        ws.cell(row=r, column=base_col + 2).value = row["format"]
        ws.cell(row=r, column=base_col + 3).value = row["taille"]
        ws.cell(row=r, column=base_col + 4).value = row["commentaire"]
        ws.cell(row=r, column=base_col + 5).value = row["exemple"]

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="Genere un mapping de champs COBOL <-> DDL Oracle Rdb.")
    parser.add_argument("--cob", required=True, help="Fichier .COB source")
    parser.add_argument("--ddl", action="append", required=True, help="Fichier(s) d'extrait DDL Rdb (repetable)")
    parser.add_argument("--table", action="append", required=True, help="Nom(s) de table a chercher, par ordre de priorite (repetable)")
    parser.add_argument("--prefix", default=None, help="Prefixe COBOL a retirer du nom de champ (ex: ENR-REFCDE-)")
    parser.add_argument(
        "--section",
        choices=["working-storage", "linkage", "all"],
        default="working-storage",
        help="Section COBOL a analyser (defaut : working-storage)",
    )
    parser.add_argument("--xlsx", help="Chemin du xlsx a mettre a jour (sinon impression CSV sur stdout)")
    parser.add_argument("--sheet-col", default="Q", help="Colonne de depart dans le xlsx (ex: Q)")
    parser.add_argument("--sheet-start-row", type=int, default=3, help="Ligne de depart des donnees")
    parser.add_argument("--sheet-header-row", type=int, default=2, help="Ligne des en-tetes")
    args = parser.parse_args()

    rows = build_rows(args.cob, args.ddl, args.table, args.prefix, args.section)

    if args.xlsx:
        write_xlsx(args.xlsx, rows, args.sheet_col, args.sheet_start_row, args.sheet_header_row)
        print(f"{len(rows)} champs ecrits dans {args.xlsx} a partir de la colonne {args.sheet_col}.")
    else:
        write_csv(rows)


if __name__ == "__main__":
    main()
